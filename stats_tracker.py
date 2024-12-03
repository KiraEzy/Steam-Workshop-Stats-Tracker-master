import requests
import time
import threading
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
from datetime import datetime
import yaml

# Initialize global variables
API_KEY = ''
WORKSHOP_ITEM_ID = ''
interval = 60
excel_file_path = 'workshop_data.xlsx'
tracking_data = []
labels = {}
fields = []
canvas = None
line_subs = None
ax = None
fig = None

def open_config_ui():
    # Create the configuration window as the main window
    config_window = tk.Tk()
    config_window.title("Configuration")

    # Configure grid to make column 1 (Entry widgets) expand
    config_window.columnconfigure(0, weight=0)
    config_window.columnconfigure(1, weight=1)

    # Create labels and entry widgets for each configuration parameter
    tk.Label(config_window, text="API Key:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
    api_key_entry = tk.Entry(config_window)
    api_key_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

    tk.Label(config_window, text="Workshop Item ID:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
    workshop_id_entry = tk.Entry(config_window)
    workshop_id_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

    tk.Label(config_window, text="Interval (seconds):").grid(row=2, column=0, sticky='e', padx=5, pady=5)
    interval_entry = tk.Entry(config_window)
    interval_entry.grid(row=2, column=1, sticky='ew', padx=5, pady=5)

    tk.Label(config_window, text="Excel File Path:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
    excel_path_entry = tk.Entry(config_window)
    excel_path_entry.grid(row=3, column=1, sticky='ew', padx=5, pady=5)

    # Load existing configuration if available
    try:
        with open('config.yaml', 'r') as config_file:
            existing_config = yaml.safe_load(config_file) or {}
            api_key_entry.insert(0, existing_config.get('api_key', ''))
            workshop_id_entry.insert(0, existing_config.get('workshop_item_id', ''))
            interval_entry.insert(0, str(existing_config.get('interval', '60')))
            excel_path_entry.insert(0, existing_config.get('excel_file_path', 'workshop_data.xlsx'))
    except FileNotFoundError:
        # No existing config, fields remain empty
        print("config.yaml not found. Starting with empty configuration.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load configuration: {e}")
        print(f"Error loading configuration: {e}")

    def apply_settings():
        # Get values from entry widgets
        api_key = api_key_entry.get().strip()
        workshop_id = workshop_id_entry.get().strip()
        interval_value = interval_entry.get().strip()
        excel_path = excel_path_entry.get().strip()

        if not api_key or not workshop_id:
            messagebox.showerror("Error", "API Key and Workshop Item ID are required.")
            return

        # Validate interval
        try:
            interval_value = int(interval_value)
        except ValueError:
            messagebox.showerror("Error", "Interval must be an integer.")
            return

        # Save to config.yaml
        new_config = {
            'api_key': api_key,
            'workshop_item_id': workshop_id,
            'interval': interval_value,
            'excel_file_path': excel_path
        }
        try:
            with open('config.yaml', 'w') as config_file:
                yaml.safe_dump(new_config, config_file)
            messagebox.showinfo("Success", "Configuration applied successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save configuration: {e}")
            print(f"Error saving configuration: {e}")

    def start_program():
        # Apply settings first
        apply_settings()
        # Close the configuration window
        config_window.destroy()
        # Start the main program
        main_program()

    # Create a frame for the buttons
    button_frame = tk.Frame(config_window)
    button_frame.grid(row=4, column=0, columnspan=2, pady=10, padx=5, sticky='e')

    # Place the buttons inside the frame
    apply_button = tk.Button(button_frame, text="Apply", command=apply_settings)
    apply_button.pack(side='right', padx=5)

    start_button = tk.Button(button_frame, text="Start", command=start_program)
    start_button.pack(side='right')

    # Start the Tkinter event loop
    config_window.mainloop()

def main_program():
    # Load configuration
    try:
        with open('config.yaml', 'r') as config_file:
            config = yaml.safe_load(config_file) or {}
    except FileNotFoundError:
        messagebox.showerror("Error", "Configuration file not found. Please apply settings first.")
        return
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load configuration: {e}")
        print(f"Error loading configuration: {e}")
        return

    # Access configuration parameters
    global API_KEY, WORKSHOP_ITEM_ID, interval, excel_file_path
    API_KEY = config.get('api_key', '')
    WORKSHOP_ITEM_ID = config.get('workshop_item_id', '')
    interval = config.get('interval', 60)
    excel_file_path = config.get('excel_file_path', 'workshop_data.xlsx')

    if not API_KEY or not WORKSHOP_ITEM_ID:
        messagebox.showerror("Error", "API Key and Workshop Item ID are required. Please apply settings first.")
        return

    # Create the main UI window
    global root
    root = tk.Tk()
    root.title("Steam Workshop Item Tracker")

    # Initialize main UI components
    initialize_main_ui()

    # Start the Tkinter event loop for the main UI
    root.mainloop()

def initialize_main_ui():
    global labels, fields, canvas, line_subs, ax, fig

    labels = {}
    fields = [
        'timestamp', 'title', 'views', 'subscriptions', 'favorites',
        'lifetime_subscriptions', 'lifetime_favorited'
    ]
    for idx, field in enumerate(fields):
        tk.Label(root, text=field.replace('_', ' ').title() + ":").grid(row=idx, column=0, sticky='e')
        labels[field] = tk.Label(root, text="")
        labels[field].grid(row=idx, column=1, sticky='w')

    # Initialize matplotlib figure
    fig, ax = plt.subplots(figsize=(8, 5))

    # Plot line for subscriptions
    line_subs, = ax.plot([], [], 's-', label='Subscriptions', color='green')

    ax.set_xlabel('Time')
    ax.set_ylabel('Count')
    ax.set_title('Workshop Item Metrics Over Time')
    ax.legend()

    # Embed the figure in the Tkinter window
    global canvas
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().grid(row=len(fields), column=0, columnspan=2)

    # Start tracking in a separate thread
    tracking_thread = threading.Thread(target=track_item)
    tracking_thread.daemon = True
    tracking_thread.start()

def fetch_workshop_item_details():
    url = 'https://api.steampowered.com/IPublishedFileService/GetDetails/v1/'
    params = {
        'key': API_KEY,
        'publishedfileids[0]': WORKSHOP_ITEM_ID
    }
    try:
        response = requests.get(url, params=params)
        if response.status_code == 200:
            data = response.json()
            item_details = data['response']['publishedfiledetails'][0]
            return item_details
        else:
            print(f"Error fetching data: {response.status_code}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
        return None

def save_to_excel(data):
    # Get current month and year
    current_month = datetime.now().strftime('%m-%Y')
    # Update the excel file path with the current month
    excel_file = excel_file_path.replace('.xlsx', f'_{current_month}.xlsx')

    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tracking Data"
        ws.append([
            'Timestamp', 'Title', 'Views', 'Subscriptions', 'Favorites',
            'Lifetime Subscriptions', 'Lifetime Favorites'
        ])
        wb.save(excel_file)
    else:
        wb = load_workbook(excel_file)
        ws = wb["Tracking Data"]

    latest_data = data[-1]
    ws.append([
        latest_data['timestamp'],
        latest_data['title'],
        latest_data['views'],
        latest_data['subscriptions'],
        latest_data['favorites'],
        latest_data['lifetime_subscriptions'],
        latest_data['lifetime_favorited'],
    ])
    wb.save(excel_file)

def update_ui(data):
    for field in fields:
        labels[field].config(text=str(data[field]))

    # Update the graph data
    times = [point['timestamp'] for point in tracking_data]
    subscriptions = [point['subscriptions'] for point in tracking_data]

    # Convert time strings to matplotlib date format
    time_nums = [mdates.datestr2num(t) for t in times]

    # Update data for the line
    line_subs.set_data(time_nums, subscriptions)

    # Adjust the axes
    ax.relim()
    ax.autoscale_view()

    # Format the x-axis to show time properly
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    fig.autofmt_xdate()

    canvas.draw()
    root.update_idletasks()

def track_item():
    while True:
        item_details = fetch_workshop_item_details()
        if item_details:
            data_point = {
                'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
                'title': item_details.get('title', 'N/A'),
                'views': int(item_details.get('views', 0)),
                'subscriptions': int(item_details.get('subscriptions', 0)),
                'favorites': int(item_details.get('favorited', 0)),
                'lifetime_subscriptions': int(item_details.get('lifetime_subscriptions', 0)),
                'lifetime_favorited': int(item_details.get('lifetime_favorited', 0)),
            }
            tracking_data.append(data_point)
            # Schedule the UI update on the main thread
            root.after(0, update_ui, data_point)
            print(f"Data fetched at {data_point['timestamp']}")
            save_to_excel(tracking_data)
        else:
            print("Failed to fetch data.")
        time.sleep(interval)

if __name__ == "__main__":
    # Start with the configuration UI
    open_config_ui()